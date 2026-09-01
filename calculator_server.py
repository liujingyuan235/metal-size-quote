from __future__ import annotations

import argparse
import contextlib
import importlib.util
import io
import json
import mimetypes
import os
import sys
import threading
import traceback
import uuid
from datetime import datetime
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import unquote, urlparse

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
ENGINE_FILE = BASE_DIR / 'backCal4cmd20251209.py'
WEB_FILE = BASE_DIR / 'calculator.html'
OUTPUT_ROOT = BASE_DIR / 'runtime'
CALCULATION_LOCK = threading.Lock()

os.environ.setdefault('MPLBACKEND', 'Agg')


def load_engine():
    spec = importlib.util.spec_from_file_location('back_calculation_engine', ENGINE_FILE)
    if spec is None or spec.loader is None:
        raise RuntimeError(f'无法加载反算代码：{ENGINE_FILE}')
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


ENGINE = load_engine()


def number(value: Any, name: str, minimum: float | None = None) -> float:
    try:
        result = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f'{name} 必须是数字') from exc
    if minimum is not None and result < minimum:
        raise ValueError(f'{name} 不能小于 {minimum}')
    return result


def integer(value: Any, name: str, minimum: int = 0) -> int:
    result = int(number(value, name, minimum))
    if result != float(value):
        raise ValueError(f'{name} 必须是整数')
    return result


def normalize_payload(raw: dict[str, Any]) -> dict[str, Any]:
    raw_orders = raw.get('orderForm')
    if not isinstance(raw_orders, list) or not raw_orders:
        raise ValueError('至少需要录入一条订单尺寸')
    if len(raw_orders) > 200:
        raise ValueError('单次最多计算 200 种订单尺寸')

    orders = []
    total_quantity = 0
    for index, raw_order in enumerate(raw_orders, start=1):
        if not isinstance(raw_order, dict):
            raise ValueError(f'第 {index} 行订单格式错误')
        quantity = integer(raw_order.get('num'), f'第 {index} 行数量', 1)
        total_quantity += quantity
        orders.append({
            'id': str(raw_order.get('id') or f'ORD-{index:03d}'),
            'orderNo': str(raw_order.get('orderNo') or f'ORD-{index:03d}'),
            'sizelong': integer(raw_order.get('sizelong'), f'第 {index} 行长度', 1),
            'sizewide': integer(raw_order.get('sizewide'), f'第 {index} 行宽度', 1),
            'direction': integer(raw_order.get('direction', 0), f'第 {index} 行纹路方向', 0),
            'num': quantity,
            'producer': str(raw_order.get('producer') or ''),
        })
    if total_quantity > 5000:
        raise ValueError('单次订单总数量不能超过 5000 件')

    min_long = integer(raw.get('min_long', 3000), '最小原板长度', 1)
    max_long = integer(raw.get('max_long', 4000), '最大原板长度', min_long)
    min_wide = integer(raw.get('min_wide', 1000), '最小原板宽度', 1)
    max_wide = integer(raw.get('max_wide', 1500), '最大原板宽度', min_wide)

    density = number(raw.get('density', 7.85), '材料密度', 0.01)
    return {
        'cutId': str(raw.get('cutId') or f'WEB-{datetime.now():%Y%m%d%H%M%S}'),
        'max_long': max_long,
        'min_long': min_long,
        'max_wide': max_wide,
        'min_wide': min_wide,
        'max_weight': number(raw.get('max_weight', 7.84), '单板最大重量', 0.01),
        'min_rate': number(raw.get('min_rate', 0.85), '最低利用率', 0),
        'kerf': integer(raw.get('kerf', 5), '刀口厚度', 0),
        'tresh': integer(raw.get('tresh', 50), '余料门限', 0),
        'max_specs': integer(raw.get('max_specs', 20), '最多规格数', 1),
        'cut_mode': integer(raw.get('cut_mode', 1), '切割模式', 1),
        'priority': integer(raw.get('priority', 1), '排序优先级', 1),
        'left_merge': integer(raw.get('left_merge', 1), '余料处理方式', 1),
        'printNo': bool(raw.get('printNo', False)),
        'thickness': number(raw.get('thickness', 20), '材料厚度', 0.1),
        'density': density,
        'forceParam': '0',
        'otherParam': {
            'density': density,
            'unitPrice': number(raw.get('unitPrice', 6.5), '材料单价', 0),
            'processingFee': number(raw.get('processingFee', 0), '加工费', 0),
        },
        'orderForm': orders,
    }


def parse_order_workbook(content: bytes) -> dict[str, Any]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f'无法读取 Excel 文件：{exc}') from exc

    aliases = {
        'sizelong': {'长', '长度', '长(mm)', '长度(mm)', 'sizelong', 'length'},
        'sizewide': {'宽', '宽度', '宽(mm)', '宽度(mm)', 'sizewide', 'width'},
        'thickness': {'厚', '厚度', '厚(mm)', '厚度(mm)', 'thickness'},
        'direction': {'纹路', '方向', '纹路方向', 'direction'},
        'num': {'数量', '件数', 'num', 'quantity'},
        'orderNo': {'订单编号', '订单号', '编号', 'orderNo', 'id'},
        'producer': {'厂家', '供应商', 'producer'},
    }

    worksheet = workbook.active
    header_row = None
    column_map: dict[str, int] = {}
    for row_number, row in enumerate(worksheet.iter_rows(min_row=1, max_row=min(20, worksheet.max_row), values_only=True), start=1):
        candidate: dict[str, int] = {}
        for column_index, cell_value in enumerate(row):
            header = str(cell_value or '').strip().replace(' ', '').lower()
            for field, field_aliases in aliases.items():
                normalized_aliases = {alias.replace(' ', '').lower() for alias in field_aliases}
                if header in normalized_aliases:
                    candidate[field] = column_index
                    break
        if {'sizelong', 'sizewide', 'num'}.issubset(candidate):
            header_row = row_number
            column_map = candidate
            break

    if header_row is None:
        raise ValueError('未找到“长、宽、数量”表头，请检查 Excel 第一张工作表')

    orders = []
    thickness_values: set[float] = set()
    warnings = []
    for excel_row, row in enumerate(worksheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not any(value is not None and str(value).strip() for value in row):
            continue
        try:
            length = integer(row[column_map['sizelong']], f'Excel 第 {excel_row} 行长度', 1)
            width = integer(row[column_map['sizewide']], f'Excel 第 {excel_row} 行宽度', 1)
            quantity = integer(row[column_map['num']], f'Excel 第 {excel_row} 行数量', 1)
            direction = integer(row[column_map['direction']], f'Excel 第 {excel_row} 行纹路', 0) if 'direction' in column_map and row[column_map['direction']] not in (None, '') else 0
            thickness = number(row[column_map['thickness']], f'Excel 第 {excel_row} 行厚度', 0.1) if 'thickness' in column_map and row[column_map['thickness']] not in (None, '') else None
            if direction not in (0, 1, 2):
                raise ValueError(f'Excel 第 {excel_row} 行纹路只能为 0、1 或 2')
            if thickness is not None:
                thickness_values.add(thickness)
            order_no = str(row[column_map['orderNo']]).strip() if 'orderNo' in column_map and row[column_map['orderNo']] not in (None, '') else f'EXCEL-{len(orders) + 1:03d}'
            producer = str(row[column_map['producer']]).strip() if 'producer' in column_map and row[column_map['producer']] not in (None, '') else ''
            orders.append({
                'id': f'EXCEL-{len(orders) + 1:03d}',
                'orderNo': order_no,
                'sizelong': length,
                'sizewide': width,
                'num': quantity,
                'direction': direction,
                'producer': producer,
                'thickness': thickness,
            })
        except ValueError as exc:
            warnings.append(str(exc))

    if not orders:
        raise ValueError('Excel 中没有可导入的有效订单行')
    if len(thickness_values) > 1:
        warnings.append('文件中存在多个厚度；反算算法一次只能处理同一厚度，当前采用第一种厚度')

    workbook.close()
    return {
        'sheetName': worksheet.title,
        'headerRow': header_row,
        'rowCount': len(orders),
        'orders': orders,
        'thickness': next(iter(thickness_values), None),
        'thicknessValues': sorted(thickness_values),
        'mappedColumns': list(column_map.keys()),
        'warnings': warnings,
    }


def execute_calculation(raw_payload: dict[str, Any]) -> dict[str, Any]:
    payload = normalize_payload(raw_payload)
    run_id = f'{datetime.now():%Y%m%d-%H%M%S}-{uuid.uuid4().hex[:6]}'
    run_dir = OUTPUT_ROOT / run_id
    image_root = run_dir / 'images'
    input_file = run_dir / 'input.json'
    result_file = run_dir / 'cut_results.json'
    run_dir.mkdir(parents=True, exist_ok=True)
    image_root.mkdir(parents=True, exist_ok=True)
    input_file.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')

    stdout_buffer = io.StringIO()
    with CALCULATION_LOCK:
        old_argv = sys.argv[:]
        old_image_root = os.environ.get('BACKCAL_IMAGE_ROOT')
        old_result_file = os.environ.get('BACKCAL_RESULT_FILE')
        try:
            ENGINE.msg = '正常运行结束'
            os.environ['BACKCAL_IMAGE_ROOT'] = str(image_root)
            os.environ['BACKCAL_RESULT_FILE'] = str(result_file)
            sys.argv = [str(ENGINE_FILE), '--input-file', str(input_file)]
            with contextlib.redirect_stdout(stdout_buffer), contextlib.redirect_stderr(stdout_buffer):
                ENGINE.main()
        finally:
            sys.argv = old_argv
            if old_image_root is None:
                os.environ.pop('BACKCAL_IMAGE_ROOT', None)
            else:
                os.environ['BACKCAL_IMAGE_ROOT'] = old_image_root
            if old_result_file is None:
                os.environ.pop('BACKCAL_RESULT_FILE', None)
            else:
                os.environ['BACKCAL_RESULT_FILE'] = old_result_file

    if not result_file.exists():
        raise RuntimeError('反算程序没有生成结果文件')

    result = json.loads(result_file.read_text(encoding='utf-8'))
    density = payload['otherParam']['density']
    unit_price = payload['otherParam']['unitPrice']
    processing_fee = payload['otherParam']['processingFee']
    thickness = payload['thickness']
    panels = result.get('result') or []

    total_area = 0.0
    for panel in panels:
        total_area += float(panel.get('sizelong', 0)) * float(panel.get('sizewide', 0))
        relative_image = str(panel.get('img') or '').replace('\\', '/')
        panel['imageUrl'] = f'/outputs/{run_id}/images/{relative_image}' if relative_image else None

    total_weight = total_area * thickness * density / 1_000_000
    material_amount = total_weight * unit_price
    result['runId'] = run_id
    result['summary'] = {
        'panelCount': len(panels),
        'totalAreaM2': round(total_area / 1_000_000, 4),
        'totalWeightKg': round(total_weight, 3),
        'materialAmount': round(material_amount, 2),
        'processingFee': round(processing_fee, 2),
        'totalPrice': round(material_amount + processing_fee, 2),
        'density': density,
        'unitPrice': unit_price,
        'thickness': thickness,
    }
    result['engineLog'] = stdout_buffer.getvalue()[-12000:]
    return result


class CalculatorHandler(BaseHTTPRequestHandler):
    server_version = 'BackCalculationWeb/1.0'

    def log_message(self, format_string: str, *args: Any) -> None:
        print(f'[{self.log_date_time_string()}] {format_string % args}')

    def send_json(self, data: dict[str, Any], status: HTTPStatus = HTTPStatus.OK) -> None:
        body = json.dumps(data, ensure_ascii=False).encode('utf-8')
        self.send_response(status)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Content-Length', str(len(body)))
        self.send_header('Cache-Control', 'no-store')
        self.end_headers()
        self.wfile.write(body)

    def send_file(self, file_path: Path) -> None:
        if not file_path.exists() or not file_path.is_file():
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        body = file_path.read_bytes()
        content_type = mimetypes.guess_type(str(file_path))[0] or 'application/octet-stream'
        self.send_response(HTTPStatus.OK)
        self.send_header('Content-Type', content_type)
        self.send_header('Content-Length', str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path in ('/', '/calculator.html'):
            self.send_file(WEB_FILE)
            return
        if parsed.path == '/api/health':
            self.send_json({'ok': True, 'engine': ENGINE_FILE.name})
            return
        if parsed.path.startswith('/outputs/'):
            relative_path = unquote(parsed.path.removeprefix('/outputs/'))
            target = (OUTPUT_ROOT / relative_path).resolve()
            if not target.is_relative_to(OUTPUT_ROOT.resolve()):
                self.send_error(HTTPStatus.FORBIDDEN)
                return
            self.send_file(target)
            return
        self.send_error(HTTPStatus.NOT_FOUND)

    def do_POST(self) -> None:
        request_path = urlparse(self.path).path
        if request_path not in ('/api/calculate', '/api/import-excel'):
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        try:
            content_length = int(self.headers.get('Content-Length', '0'))
            limit = 10 * 1024 * 1024 if request_path == '/api/import-excel' else 2 * 1024 * 1024
            if content_length <= 0 or content_length > limit:
                raise ValueError(f'请求内容为空或超过 {limit // 1024 // 1024}MB 限制')
            content = self.rfile.read(content_length)
            if request_path == '/api/import-excel':
                self.send_json({'ok': True, 'data': parse_order_workbook(content)})
                return
            payload = json.loads(content.decode('utf-8'))
            if not isinstance(payload, dict):
                raise ValueError('请求必须是 JSON 对象')
            self.send_json({'ok': True, 'data': execute_calculation(payload)})
        except ValueError as exc:
            self.send_json({'ok': False, 'message': str(exc)}, HTTPStatus.BAD_REQUEST)
        except Exception as exc:
            traceback.print_exc()
            self.send_json({'ok': False, 'message': f'反算失败：{exc}'}, HTTPStatus.INTERNAL_SERVER_ERROR)


def main() -> None:
    parser = argparse.ArgumentParser(description='金属订单反算报价网页服务')
    parser.add_argument('--host', default=os.environ.get('HOST', '127.0.0.1'))
    parser.add_argument('--port', type=int, default=int(os.environ.get('PORT', '8090')))
    args = parser.parse_args()

    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    server = ThreadingHTTPServer((args.host, args.port), CalculatorHandler)
    print(f'反算报价链接：http://{args.host}:{args.port}/')
    print(f'反算引擎：{ENGINE_FILE.name}')
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print('服务已停止')
    finally:
        server.server_close()


if __name__ == '__main__':
    main()
